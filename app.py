import io
import streamlit as st
import pandas as pd
import plotly.express as px
import geopandas as gpd
import plotly.graph_objects as go
from streamlit_option_menu import option_menu
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- PAGE CONFIG ---
st.set_page_config(
    page_title="DSS Bansos Jawa Timur",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- GLOBAL CSS ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    :root {
        --primary: #1B6CA8;
        --primary-light: #EBF4FF;
        --accent: #0EA5E9;
        --success: #10B981;
        --danger: #EF4444;
        --warning: #F59E0B;
        --surface: #FFFFFF;
        --surface-2: #F8FAFC;
        --border: #E2E8F0;
        --text-primary: #0F172A;
        --text-secondary: #64748B;
        --text-muted: #94A3B8;
        --shadow-sm: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
        --shadow-md: 0 4px 12px rgba(0,0,0,0.08), 0 2px 4px rgba(0,0,0,0.04);
        --radius: 10px;
    }
    html, body, [class*="css"] {
        font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif !important;
        color: var(--text-primary);
    }
    .main .block-container { padding: 2rem 2.5rem 3rem !important; max-width: 1400px; }

    /* SIDEBAR */
    [data-testid="stSidebar"] { background: #0F172A !important; border-right: none !important; }
    [data-testid="stSidebar"] * { color: #CBD5E1 !important; }
    [data-testid="stSidebar"] .stMarkdown p, [data-testid="stSidebar"] label {
        color: #94A3B8 !important; font-size: 12px !important; font-weight: 500 !important;
        letter-spacing: 0.05em !important; text-transform: uppercase !important;
    }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 { color: #F1F5F9 !important; }
    [data-testid="stSidebar"] hr { border-color: #1E293B !important; margin: 1rem 0 !important; }
    [data-testid="stSidebar"] [data-baseweb="select"] > div,
    [data-testid="stSidebar"] [data-baseweb="input"] input {
        background: #1E293B !important; border-color: #334155 !important;
        color: #E2E8F0 !important; border-radius: 8px !important;
    }
    [data-testid="stSidebar"] [data-baseweb="slider"] [role="slider"] { background: var(--accent) !important; }
    [data-testid="stSidebar"] .stAlert {
        background: #1E293B !important; border: 1px solid #334155 !important;
        border-radius: 8px !important; color: #94A3B8 !important; font-size: 12px !important;
    }
    [data-testid="stSidebar"] .nav-link { border-radius: 8px !important; margin: 2px 0 !important; font-weight: 500 !important; font-size: 13.5px !important; color: #94A3B8 !important; }
    [data-testid="stSidebar"] .nav-link:hover { background: #1E293B !important; color: #F1F5F9 !important; }
    [data-testid="stSidebar"] .nav-link-selected { background: var(--primary) !important; color: #FFFFFF !important; }

    /* MAIN */
    h1 { font-size: 1.75rem !important; font-weight: 700 !important; letter-spacing: -0.02em !important; color: var(--text-primary) !important; margin-bottom: 0.25rem !important; }
    h2 { font-size: 1.2rem !important; font-weight: 600 !important; color: var(--text-primary) !important; }
    h3 { font-size: 1.05rem !important; font-weight: 600 !important; color: var(--text-primary) !important; }

    [data-testid="metric-container"] {
        background: var(--surface) !important; border: 1px solid var(--border) !important;
        border-radius: var(--radius) !important; padding: 1rem 1.25rem !important;
        box-shadow: var(--shadow-sm) !important; transition: box-shadow 0.2s ease !important;
    }
    [data-testid="metric-container"]:hover { box-shadow: var(--shadow-md) !important; }
    [data-testid="stMetricLabel"] { font-size: 11px !important; font-weight: 600 !important; color: var(--text-muted) !important; text-transform: uppercase !important; letter-spacing: 0.06em !important; }
    [data-testid="stMetricValue"] { font-size: 1.5rem !important; font-weight: 700 !important; color: var(--text-primary) !important; letter-spacing: -0.02em !important; font-family: 'DM Mono', monospace !important; }
    [data-testid="stMetricDelta"] { font-size: 12px !important; font-weight: 500 !important; }

    [data-testid="stVerticalBlockBorderWrapper"] {
        border: 1px solid var(--border) !important; border-radius: var(--radius) !important;
        background: var(--surface) !important; box-shadow: var(--shadow-sm) !important;
        overflow: hidden !important; transition: box-shadow 0.2s ease !important;
    }
    [data-testid="stVerticalBlockBorderWrapper"]:hover { box-shadow: var(--shadow-md) !important; }

    [data-baseweb="select"] > div {
        border-radius: 8px !important; border-color: var(--border) !important;
        background: var(--surface) !important; font-size: 14px !important;
        transition: border-color 0.2s ease, box-shadow 0.2s ease !important;
    }
    [data-baseweb="select"] > div:focus-within {
        border-color: var(--primary) !important; box-shadow: 0 0 0 3px rgba(27,108,168,0.12) !important;
    }
    [data-testid="stDataFrame"] { border: 1px solid var(--border) !important; border-radius: var(--radius) !important; overflow: hidden !important; }
    .stAlert { border-radius: 8px !important; border: none !important; font-size: 13px !important; }
    .stAlert [data-testid="stMarkdownContainer"] p { font-size: 13px !important; color: var(--text-secondary) !important; }

    [data-testid="baseButton-primary"] {
        background: var(--primary) !important; color: white !important; border: none !important;
        border-radius: 8px !important; font-weight: 600 !important; font-size: 13.5px !important;
        letter-spacing: 0.01em !important; padding: 0.5rem 1rem !important;
        transition: all 0.2s ease !important; box-shadow: 0 2px 8px rgba(27,108,168,0.25) !important;
    }
    [data-testid="baseButton-primary"]:hover {
        background: #155E8E !important; box-shadow: 0 4px 12px rgba(27,108,168,0.35) !important; transform: translateY(-1px) !important;
    }
    [data-testid="baseButton-secondary"] { border-radius: 8px !important; font-size: 13px !important; border-color: var(--border) !important; }

    hr { border-color: var(--border) !important; margin: 1.5rem 0 !important; }
    [data-baseweb="input"] input { border-radius: 8px !important; font-family: 'DM Mono', monospace !important; font-size: 13px !important; }
    [data-testid="stSlider"] [role="slider"] { background: var(--primary) !important; }
    .js-plotly-plot { border-radius: var(--radius) !important; }

    /* TABS */
    [data-testid="stTabs"] [data-baseweb="tab-list"] { gap: 4px; border-bottom: 2px solid var(--border) !important; background: transparent !important; }
    [data-testid="stTabs"] [data-baseweb="tab"] {
        border-radius: 6px 6px 0 0 !important; font-size: 13px !important; font-weight: 600 !important;
        color: var(--text-muted) !important; padding: 8px 16px !important;
        background: transparent !important; border: none !important;
    }
    [data-testid="stTabs"] [aria-selected="true"] { color: var(--primary) !important; border-bottom: 2px solid var(--primary) !important; }

    /* CUSTOM CLASSES */
    .section-label { font-size: 11px; font-weight: 700; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 0.75rem; display: block; }
    .page-subtitle { font-size: 14px; color: var(--text-secondary); margin-top: -0.5rem; margin-bottom: 1.5rem; font-weight: 400; }
    .header-tag { display: inline-block; background: var(--primary-light); color: var(--primary); font-size: 11px; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase; padding: 3px 10px; border-radius: 20px; margin-bottom: 0.5rem; }
    .card-title { font-size: 13px; font-weight: 700; color: var(--text-primary); margin-bottom: 2px; }
    .card-subtitle { font-size: 12px; color: var(--text-muted); margin-bottom: 0.75rem; font-weight: 400; }
    .brand-block { padding: 1rem 0 0.5rem 0; }
    .brand-title { font-size: 11px; font-weight: 700; color: #94A3B8; letter-spacing: 0.06em; text-transform: uppercase; line-height: 1.5; }
    .brand-sub { font-size: 10px; color: #64748B; font-weight: 400; line-height: 1.4; }

    /* SIDEBAR EXPANDER */
    [data-testid="stSidebar"] [data-testid="stExpander"] { border: 1px solid #1E293B !important; border-radius: 8px !important; background: #0D1929 !important; overflow: hidden !important; }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary { background: transparent !important; color: #64748B !important; font-size: 11px !important; font-weight: 700 !important; letter-spacing: 0.07em !important; text-transform: uppercase !important; padding: 10px 12px !important; }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary:hover { color: #94A3B8 !important; background: #1E293B !important; }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary svg { fill: #475569 !important; }
    [data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stExpanderDetails"] { background: #0D1929 !important; padding: 0 12px 12px !important; border-top: 1px solid #1E293B !important; }

    .src-item { display: flex; gap: 10px; align-items: flex-start; padding: 8px 0; border-bottom: 1px solid #1E293B; }
    .src-item:last-child { border-bottom: none; padding-bottom: 0; }
    .src-badge { flex-shrink: 0; width: 20px; height: 20px; background: #1B6CA822; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 10px; font-weight: 700; color: #4A9DD4; margin-top: 1px; }
    .src-content { flex: 1; }
    .src-label { font-size: 10px; font-weight: 700; color: #94A3B8; text-transform: uppercase; letter-spacing: .06em; margin-bottom: 2px; }
    .src-desc { font-size: 11px; color: #64748B; line-height: 1.5; }
    .src-desc i { color: #475569; font-style: italic; }

    .scenario-badge { display: inline-flex; align-items: center; gap: 6px; padding: 4px 10px; border-radius: 20px; font-size: 11px; font-weight: 600; margin: 2px; }
    .scenario-dot { width: 8px; height: 8px; border-radius: 50%; display: inline-block; }
</style>
""", unsafe_allow_html=True)


# ==============================================================================
# DATA & ENGINE (CACHED)
# ==============================================================================
@st.cache_data
def load_data():
    df = pd.read_csv("bansos_jatim.csv", sep=";")
    df.columns = df.columns.str.strip()
    return df

@st.cache_data
def load_geodata():
    return gpd.read_file("jawa-timur-simplified-topo.json")

@st.cache_data
def hitung_optimasi(df_input: pd.DataFrame, kolom_bansos: str, total_pagu: int, toleransi: float):
    df = df_input.copy()
    df['Alokasi_Murni'] = df['Porsi_Miskin'] * total_pagu
    df['Batas_Bawah'] = df[kolom_bansos] * (1 - toleransi)
    df['Batas_Atas']  = df[kolom_bansos] * (1 + toleransi)

    def _bounded(row):
        if row['Alokasi_Murni'] < row['Batas_Bawah']:
            return row['Batas_Bawah']
        elif row['Alokasi_Murni'] > row['Batas_Atas']:
            return row['Batas_Atas']
        return row['Alokasi_Murni']

    df['Alokasi_Rekomendasi'] = df.apply(_bounded, axis=1)
    df['Gap'] = df['Alokasi_Rekomendasi'] - df[kolom_bansos]
    return df

def buat_excel(df: pd.DataFrame, kolom_bansos: str) -> bytes:
    cols = ['Kabupaten_Kota', kolom_bansos, 'Alokasi_Murni', 'Alokasi_Rekomendasi', 'Gap']
    df_exp = df[cols].copy()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hasil Optimasi"

    header_fill = PatternFill("solid", fgColor="1B6CA8")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Kabupaten / Kota", kolom_bansos, "Alokasi Murni", "Rekomendasi", "Gap"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    green_fill = PatternFill("solid", fgColor="D1FAE5")
    red_fill   = PatternFill("solid", fgColor="FEE2E2")
    green_font = Font(bold=True, color="065F46", name="Calibri", size=10)
    red_font   = Font(bold=True, color="991B1B", name="Calibri", size=10)
    normal_font= Font(name="Calibri", size=10)

    for ri, row in enumerate(df_exp.itertuples(index=False), 2):
        for ci, val in enumerate(list(row), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            if ci == 1:
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
                if ci == 5:
                    if isinstance(val, (int, float)) and val < 0:
                        cell.fill = red_fill; cell.font = red_font
                    elif isinstance(val, (int, float)) and val > 0:
                        cell.fill = green_fill; cell.font = green_font
                    else:
                        cell.font = normal_font
                else:
                    cell.font = normal_font

    for i, w in enumerate([28, 16, 16, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ==============================================================================
# INIT
# ==============================================================================
df_raw = load_data()
if 'skenario_list' not in st.session_state:
    st.session_state['skenario_list'] = []

CHART_FONT = dict(family="DM Sans, sans-serif")
WARNA_SKENARIO = ['#1B6CA8', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6']


# ==============================================================================
# SIDEBAR
# ==============================================================================
with st.sidebar:
    col_logo, col_brand = st.columns([1, 2.2])
    with col_logo:
        st.image("logo.png", use_container_width=True)
    with col_brand:
        st.markdown("""
            <div class='brand-block'>
                <div class='brand-title'>Badan Pusat Statistik</div>
                <div class='brand-sub'>Kabupaten Bojonegoro</div>
            </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    menu = option_menu(
        menu_title=None,
        options=["Sistem Pendukung Keputusan", "Metodologi"],
        icons=["map", "graph-up"],
        default_index=0,
        styles={
            "container": {"padding": "0", "background-color": "transparent"},
            "icon": {"color": "#64748B", "font-size": "14px"},
            "nav-link": {"font-size": "13.5px", "font-weight": "500", "text-align": "left",
                         "padding": "9px 12px", "margin": "2px 0", "border-radius": "8px",
                         "color": "#94A3B8", "--hover-color": "#1E293B"},
            "nav-link-selected": {"background-color": "#1B6CA8", "color": "#FFFFFF", "font-weight": "600"},
        }
    )

    st.markdown("---")
    st.markdown('<span class="section-label" style="color:#64748B;font-size:10px;letter-spacing:.08em;text-transform:uppercase;font-weight:700;">Parameter Kebijakan</span>', unsafe_allow_html=True)

    pilihan_bansos = st.selectbox("Program Bansos", ("PKH_Plus","ASPD","KE","BLT","EKS_PPKS","LKSA"))

    toleransi_persen = st.slider("Toleransi Perubahan Kuota (%)", 0, 50, 20,
                                  help="Batas maksimal perubahan kuota dari nilai eksisting (± persen)")
    toleransi = toleransi_persen / 100.0

    st.markdown("---")
    st.markdown('<span class="section-label" style="color:#64748B;font-size:10px;letter-spacing:.08em;text-transform:uppercase;font-weight:700;">Simulasi Anggaran</span>', unsafe_allow_html=True)

    total_eksisting = int(df_raw[pilihan_bansos].sum())
    total_pagu = st.number_input(f"Total Pagu {pilihan_bansos} (unit)", min_value=0, value=total_eksisting, step=100)
    st.caption(f"Nilai historis: **{total_eksisting:,}** unit")

    st.markdown("---")
    with st.expander("Sumber Data", expanded=False):
        st.markdown("""
        <div style="padding-top:4px;">
            <div class="src-item">
                <div class="src-badge">1</div>
                <div class="src-content">
                    <div class="src-label">Data Alokasi Eksisting</div>
                    <div class="src-desc">Portal <i>SAPA BANSOS</i> Provinsi Jawa Timur<br>
                    <span style="color:#334155;">Diakses 25 Desember 2025</span></div>
                </div>
            </div>
            <div class="src-item">
                <div class="src-badge">2</div>
                <div class="src-content">
                    <div class="src-label">Profil Kemiskinan</div>
                    <div class="src-desc">Publikasi resmi BPS:<br>
                    <i>"Profil Kemiskinan di Kabupaten Bojonegoro Maret 2025"</i><br>
                    <span style="color:#334155;">dan data rujukan kabupaten/kota se-Jawa Timur</span></div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)


# --- IDENTITAS KAMPUS ---
    st.markdown("<br>", unsafe_allow_html=True)
    col_logo_mat, col_teks_its = st.columns([1, 2.5])
    
    with col_logo_mat:
        st.image("logo_mat.png", use_container_width=True) 
        
    with col_teks_its:
        st.markdown("""
        <div style='padding-top: 2px; line-height: 1.3;'>
            <b style='color: #F1F5F9; font-size: 11px; letter-spacing: 0.02em;'>DEPARTMENT OF MATHEMATICS</b><br>
            <span style='color: #94A3B8; font-size: 9.5px;'>Institut Teknologi Sepuluh Nopember</span>
        </div>
        """, unsafe_allow_html=True)

# --- ENGINE (cached) ---
df = hitung_optimasi(df_raw, pilihan_bansos, total_pagu, toleransi)


# ==============================================================================
# HALAMAN 1
# ==============================================================================
if menu == "Sistem Pendukung Keputusan":

    st.markdown('<span class="header-tag">Sistem Pendukung Keputusan</span>', unsafe_allow_html=True)
    st.title("Optimasi Alokasi Bantuan Sosial Jawa Timur")
    st.markdown('<p class="page-subtitle">Pemodelan spasial berbasis porsi kemiskinan dengan batasan stabilitas kebijakan</p>', unsafe_allow_html=True)

    tab_distribusi, tab_skenario, tab_efisiensi = st.tabs([
        "Distribusi & Peta", "Perbandingan Skenario", "Analisis Efisiensi"
    ])

    # ------------------------------------------------------------------
    # TAB 1 — DISTRIBUSI & PETA
    # ------------------------------------------------------------------
    with tab_distribusi:

        st.markdown("#### Analisis Keadilan Distribusi")
        st.caption("Lima daerah dengan pergeseran kuota terbesar berdasarkan hasil optimasi proporsional.")

        col_top1, col_top2 = st.columns(2, gap="medium")

        with col_top1:
            with st.container(border=True):
                st.markdown('<p class="card-title">Prioritas Penambahan Kuota</p>'
                            '<p class="card-subtitle">Daerah dengan kekurangan alokasi relatif tertinggi</p>',
                            unsafe_allow_html=True)
                df_tambah = df.nlargest(5, 'Gap').sort_values('Gap', ascending=True)
                fig_tambah = go.Figure(go.Bar(
                    x=df_tambah['Gap'], y=df_tambah['Kabupaten_Kota'], orientation='h',
                    text=df_tambah['Gap'].apply(lambda x: f"+{x:,.0f}"),
                    textposition='inside', insidetextanchor='middle',
                    textfont=dict(color='white', size=12, family='DM Mono, monospace'),
                    marker=dict(color=df_tambah['Gap'], colorscale=[[0,'#BFDBFE'],[1,'#1B6CA8']], showscale=False),
                    width=0.55
                ))
                fig_tambah.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=0,r=10,t=5,b=0), height=260, font=CHART_FONT,
                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                    yaxis=dict(showgrid=False, zeroline=False, tickfont=dict(size=12, color='#374151'))
                )
                st.plotly_chart(fig_tambah, use_container_width=True, config={'displayModeBar': False})

        with col_top2:
            with st.container(border=True):
                st.markdown('<p class="card-title">Prioritas Rasionalisasi Kuota</p>'
                            '<p class="card-subtitle">Daerah dengan kelebihan alokasi relatif tertinggi</p>',
                            unsafe_allow_html=True)
                df_kurang = df.nsmallest(5, 'Gap').sort_values('Gap', ascending=False).copy()
                df_kurang['Abs_Gap'] = df_kurang['Gap'].abs()
                fig_kurang = go.Figure(go.Bar(
                    x=df_kurang['Abs_Gap'], y=df_kurang['Kabupaten_Kota'], orientation='h',
                    text=df_kurang['Gap'].apply(lambda x: f"{x:,.0f}"),
                    textposition='inside', insidetextanchor='middle',
                    textfont=dict(color='white', size=12, family='DM Mono, monospace'),
                    marker=dict(color=df_kurang['Abs_Gap'], colorscale=[[0,'#FECACA'],[1,'#DC2626']], showscale=False),
                    width=0.55
                ))
                fig_kurang.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=0,r=10,t=5,b=0), height=260, font=CHART_FONT,
                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                    yaxis=dict(showgrid=False, zeroline=False, tickfont=dict(size=12, color='#374151'))
                )
                st.plotly_chart(fig_kurang, use_container_width=True, config={'displayModeBar': False})

        st.markdown("---")

        # Profil daerah
        st.markdown("#### Profil Daerah")
        daftar_pilihan = ['Seluruh Jawa Timur'] + list(df['Kabupaten_Kota'].unique())
        kota_terpilih = st.selectbox("Pilih Kabupaten / Kota", daftar_pilihan)

        col_k1, col_k2, col_k3, col_k4 = st.columns(4, gap="small")
        if kota_terpilih == 'Seluruh Jawa Timur':
            with col_k1: st.metric("Penduduk Miskin (Ribu Jiwa)", f"{df['Jml_Miskin'].sum():,.2f}")
            with col_k2: st.metric("Kuota Eksisting", f"{df[pilihan_bansos].sum():,.0f}")
            with col_k3: st.metric("Rekomendasi Optimasi", f"{df['Alokasi_Rekomendasi'].sum():,.0f}")
            with col_k4:
                g = df['Gap'].sum()
                st.metric("Agregat Gap", f"{g:,.0f}", delta=f"{g:,.0f}", delta_color="normal")
        else:
            dk = df[df['Kabupaten_Kota'] == kota_terpilih].iloc[0]
            with col_k1: st.metric("Penduduk Miskin (Ribu Jiwa)", f"{dk['Jml_Miskin']:,.2f}")
            with col_k2: st.metric("Kuota Eksisting", f"{dk[pilihan_bansos]:,.0f}")
            with col_k3: st.metric("Rekomendasi Optimasi", f"{dk['Alokasi_Rekomendasi']:,.0f}")
            with col_k4:
                g = dk['Gap']
                st.metric("Gap Alokasi", f"{g:,.0f}", delta=f"{g:,.0f}", delta_color="normal")

        st.markdown("---")

        # Peta
        st.markdown("#### Peta Spasial Kesenjangan Alokasi")
        try:
            gdf = load_geodata()
            gdf['match_key'] = gdf['kabkot'].astype(str).str.upper().str.replace('KABUPATEN ', '').str.replace('KOTA ', '').str.strip()
            df['match_key'] = df['Kabupaten_Kota'].astype(str).str.upper().str.replace('KABUPATEN ', '').str.replace('KOTA ', '').str.strip()
            
            # Merge menggunakan kolom buatan 'match_key'
            gdf_m = gdf.merge(df, on='match_key', how='inner').reset_index(drop=True)

            if kota_terpilih == 'Seluruh Jawa Timur':
                lw = [0.4] * len(gdf_m); lc = ['#94A3B8'] * len(gdf_m)
                lat, lon, zoom = -7.7, 112.9, 6.5
            else:
                target_kota = gdf_m[gdf_m['Kabupaten_Kota'] == kota_terpilih]
                
                # Jika ternyata datanya masih kosong/tidak cocok
                if target_kota.empty:
                    st.warning(f"⚠️ Peringatan: Batas poligon peta untuk '{kota_terpilih}' tidak ditemukan di file GeoJSON.")
                    lw = [0.4] * len(gdf_m); lc = ['#94A3B8'] * len(gdf_m)
                    lat, lon, zoom = -7.7, 112.9, 6.5
                else:
                    lw = [3.5 if k == kota_terpilih else 0.4 for k in gdf_m['Kabupaten_Kota']]
                    lc = ['#0EA5E9' if k == kota_terpilih else '#94A3B8' for k in gdf_m['Kabupaten_Kota']]
                    bx = target_kota.total_bounds
                    lat, lon, zoom = (bx[1]+bx[3])/2, (bx[0]+bx[2])/2, 9

            fig_map = px.choropleth_mapbox(
                gdf_m, geojson=gdf_m.geometry, locations=gdf_m.index, color="Gap",
                hover_name="Kabupaten_Kota",
                hover_data={"Gap":":,.0f","Alokasi_Rekomendasi":":,.0f",pilihan_bansos:":,.0f","Jml_Miskin":":,.2f"},
                color_continuous_scale="RdBu", color_continuous_midpoint=0,
                mapbox_style="open-street-map", center={"lat":lat,"lon":lon}, zoom=zoom, opacity=0.65
            )
            fig_map.update_traces(marker=dict(line=dict(width=lw, color=lc)))
            fig_map.update_layout(
                margin=dict(t=0,l=0,r=0,b=0), height=540, paper_bgcolor='rgba(0,0,0,0)',
                coloraxis_colorbar=dict(title="Gap", thickness=14, len=0.6, tickformat=",.0f",
                                        tickfont=dict(size=11,family="DM Mono, monospace"),
                                        title_font=dict(size=11,family="DM Sans, sans-serif"))
            )
            st.plotly_chart(fig_map, use_container_width=True, config={'displayModeBar': False})
        except Exception as e:
            st.warning(f"Peta tidak dapat dimuat. Detail: {e}")

        # Grafik + Tabel
        st.markdown(f"#### Perbandingan Kuota Eksisting vs Rekomendasi — {pilihan_bansos}")
        col_chart, col_table = st.columns([6, 4], gap="medium")

        with col_chart:
            df_sorted = df.sort_values(pilihan_bansos, ascending=False)
            if kota_terpilih == 'Seluruh Jawa Timur':
                wb_color = ['#F59E0B' if 'Bojonegoro' in str(k) else '#CBD5E1' for k in df_sorted['Kabupaten_Kota']]
            else:
                wb_color = ['#1B6CA8' if k == kota_terpilih else '#F59E0B' if 'Bojonegoro' in str(k) else '#CBD5E1' for k in df_sorted['Kabupaten_Kota']]

            fig_combo = go.Figure()
            fig_combo.add_trace(go.Bar(
                x=df_sorted['Kabupaten_Kota'], y=df_sorted[pilihan_bansos],
                name='Kuota Eksisting', marker_color=wb_color,
                hovertemplate="<b>%{x}</b><br>Eksisting: %{y:,.0f}<extra></extra>"
            ))
            fig_combo.add_trace(go.Scatter(
                x=df_sorted['Kabupaten_Kota'], y=df_sorted['Alokasi_Rekomendasi'],
                name='Target Optimasi', mode='lines+markers',
                line=dict(color='#374151', width=1.5, dash='dot'),
                marker=dict(size=7, color='#EF4444', line=dict(width=1.5, color='white')),
                hovertemplate="Target: %{y:,.0f}<extra></extra>"
            ))
            fig_combo.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis_tickangle=-45, margin=dict(t=10,l=0,r=0,b=0), height=380, font=CHART_FONT,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, font=dict(size=12)),
                hovermode="x unified",
                xaxis=dict(showgrid=False, tickfont=dict(size=10)),
                yaxis=dict(showgrid=True, gridcolor='#F1F5F9', zeroline=False,
                           tickformat=",", tickfont=dict(size=11,family='DM Mono, monospace')),
                bargap=0.3
            )
            st.plotly_chart(fig_combo, use_container_width=True, config={'displayModeBar': False})

        with col_table:
            st.markdown('<p style="font-size:13px;font-weight:600;color:#374151;margin-bottom:0.5rem;">Tabel Keputusan Alokasi</p>', unsafe_allow_html=True)
            max_rek = int(df['Alokasi_Rekomendasi'].max())
            st.dataframe(
                df[['Kabupaten_Kota', pilihan_bansos, 'Alokasi_Rekomendasi', 'Gap']].style
                .format({pilihan_bansos:'{:,.0f}','Alokasi_Rekomendasi':'{:,.0f}','Gap':'{:+,.0f}'})
                .map(lambda v: 'color:#DC2626;font-weight:600' if isinstance(v,(int,float)) and v<0
                     else ('color:#059669;font-weight:600' if isinstance(v,(int,float)) and v>0 else ''),
                     subset=['Gap']),
                use_container_width=True, height=300,
                column_config={
                    "Kabupaten_Kota": st.column_config.TextColumn("Wilayah"),
                    pilihan_bansos: st.column_config.NumberColumn("Eksisting"),
                    "Alokasi_Rekomendasi": st.column_config.ProgressColumn(
                        "Rekomendasi",
                        help="Proporsi relatif terhadap nilai tertinggi",
                        format="%,.0f", min_value=0, max_value=max_rek
                    ),
                    "Gap": st.column_config.NumberColumn("Gap"),
                }
            )

            st.markdown("<br>", unsafe_allow_html=True)
            csv_data = df[['Kabupaten_Kota',pilihan_bansos,'Alokasi_Murni','Alokasi_Rekomendasi','Gap']].to_csv(index=False).encode('utf-8')
            c1, c2 = st.columns(2)
            with c1:
                if st.download_button("Unduh CSV", data=csv_data,
                                      file_name=f"Rekomendasi_{pilihan_bansos}_Jatim.csv",
                                      mime="text/csv", type="primary",
                                      use_container_width=True, key="dl_csv"):
                    st.toast("File CSV berhasil diunduh.", icon="✅")
            with c2:
                xlsx_data = buat_excel(df, pilihan_bansos)
                if st.download_button("Unduh Excel", data=xlsx_data,
                                      file_name=f"Rekomendasi_{pilihan_bansos}_Jatim.xlsx",
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                      type="primary", use_container_width=True, key="dl_xlsx"):
                    st.toast("File Excel berhasil diunduh.", icon="✅")

    # ------------------------------------------------------------------
    # TAB 2 — PERBANDINGAN SKENARIO
    # ------------------------------------------------------------------
    with tab_skenario:
        st.markdown("#### Perbandingan Skenario Toleransi")
        st.caption("Simpan beberapa konfigurasi toleransi untuk membandingkan hasilnya secara overlay. Maksimal 5 skenario.")

        col_s1, col_s2, col_s3 = st.columns([2.5, 1.5, 1], gap="small")
        with col_s1:
            nama_sk = st.text_input("Nama Skenario", value=f"Toleransi {toleransi_persen}% — Pagu {total_pagu:,}", key="nama_sk")
        with col_s2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Simpan Skenario Aktif", use_container_width=True):
                existing = [s['nama'] for s in st.session_state['skenario_list']]
                if nama_sk in existing:
                    st.warning(f"Nama '{nama_sk}' sudah digunakan.")
                elif len(st.session_state['skenario_list']) >= 5:
                    st.warning("Batas maksimal 5 skenario tercapai.")
                else:
                    st.session_state['skenario_list'].append({
                        'nama': nama_sk, 'toleransi': toleransi_persen,
                        'pagu': total_pagu, 'bansos': pilihan_bansos,
                        'data': df[['Kabupaten_Kota','Alokasi_Rekomendasi','Gap']].copy()
                    })
                    st.toast(f"Skenario '{nama_sk}' tersimpan.", icon="✅")
        with col_s3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Hapus Semua", use_container_width=True):
                st.session_state['skenario_list'] = []
                st.toast("Semua skenario dihapus.")

        if not st.session_state['skenario_list']:
            st.markdown("""
            <div style="text-align:center;padding:3rem 1rem;background:#F8FAFC;border:1px dashed #E2E8F0;border-radius:10px;margin-top:1rem;">
                <p style="color:#94A3B8;font-size:14px;margin:0;">Belum ada skenario tersimpan.<br>
                Atur parameter di sidebar lalu klik <b>Simpan Skenario Aktif</b>.</p>
            </div>""", unsafe_allow_html=True)
        else:
            # Badge labels
            badges = "".join(
                f'<span class="scenario-badge" style="background:{WARNA_SKENARIO[i%5]}18;color:{WARNA_SKENARIO[i%5]};border:1px solid {WARNA_SKENARIO[i%5]}44;">'
                f'<span class="scenario-dot" style="background:{WARNA_SKENARIO[i%5]};"></span>{s["nama"]}</span>'
                for i, s in enumerate(st.session_state['skenario_list'])
            )
            st.markdown(f'<div style="margin:.5rem 0 1rem;">{badges}</div>', unsafe_allow_html=True)

            # Overlay line chart
            fig_ov = go.Figure()
            for i, sk in enumerate(st.session_state['skenario_list']):
                w = WARNA_SKENARIO[i % 5]
                ds = sk['data'].sort_values('Kabupaten_Kota')
                fig_ov.add_trace(go.Scatter(
                    x=ds['Kabupaten_Kota'], y=ds['Alokasi_Rekomendasi'],
                    name=sk['nama'], mode='lines+markers',
                    line=dict(color=w, width=2),
                    marker=dict(size=6, color=w, line=dict(width=1.5, color='white')),
                    hovertemplate=f"<b>%{{x}}</b><br>{sk['nama']}: %{{y:,.0f}}<extra></extra>"
                ))
            fig_ov.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis_tickangle=-45, margin=dict(t=20,l=0,r=0,b=0), height=400, font=CHART_FONT,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, font=dict(size=12)),
                hovermode="x unified",
                xaxis=dict(showgrid=False, tickfont=dict(size=10)),
                yaxis=dict(showgrid=True, gridcolor='#F1F5F9', zeroline=False,
                           tickformat=",", tickfont=dict(size=11,family='DM Mono, monospace'))
            )
            st.plotly_chart(fig_ov, use_container_width=True, config={'displayModeBar': False})

            # Total absolute gap comparison
            st.markdown("##### Total Gap Absolut Antar Skenario")
            nama_list = [s['nama'] for s in st.session_state['skenario_list']]
            gap_list  = [s['data']['Gap'].abs().sum() for s in st.session_state['skenario_list']]
            fig_gb = go.Figure(go.Bar(
                x=nama_list, y=gap_list,
                marker_color=[WARNA_SKENARIO[i%5] for i in range(len(nama_list))],
                text=[f"{v:,.0f}" for v in gap_list],
                textposition='outside',
                textfont=dict(size=11,family='DM Mono, monospace',color='#374151'),
                width=0.45
            ))
            fig_gb.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(t=30,l=0,r=0,b=0), height=280, font=CHART_FONT,
                xaxis=dict(showgrid=False, tickfont=dict(size=12)),
                yaxis=dict(showgrid=True, gridcolor='#F1F5F9', zeroline=False,
                           tickformat=",", tickfont=dict(size=11,family='DM Mono, monospace')),
                bargap=0.4
            )
            st.plotly_chart(fig_gb, use_container_width=True, config={'displayModeBar': False})

    # ------------------------------------------------------------------
    # TAB 3 — ANALISIS EFISIENSI
    # ------------------------------------------------------------------
    with tab_efisiensi:
        st.markdown("#### Scatter Plot Efisiensi Alokasi")
        st.caption("Sumbu X: proporsi kemiskinan. Sumbu Y: kuota per kapita miskin. Titik jauh dari garis referensi mengindikasikan alokasi yang tidak proporsional.")

        df_eff = df.copy()
        df_eff['Kuota_Per_Kapita']       = df_eff[pilihan_bansos] / df_eff['Jml_Miskin']
        df_eff['Rekomendasi_Per_Kapita'] = df_eff['Alokasi_Rekomendasi'] / df_eff['Jml_Miskin']

        x_min, x_max = df_eff['Porsi_Miskin'].min(), df_eff['Porsi_Miskin'].max()
        x_ref = [x_min, x_max]

        col_e1, col_e2 = st.columns(2, gap="medium")

        def _ref_line(series):
            mean_x = df_eff['Porsi_Miskin'].mean()
            mean_y = series.mean()
            return [x * mean_y / mean_x for x in x_ref]

        def _scatter(y_col, title, subtitle, dot_color=None, is_colored=False):
            with st.container(border=True):
                st.markdown(f'<p class="card-title">{title}</p><p class="card-subtitle">{subtitle}</p>', unsafe_allow_html=True)
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=x_ref, y=_ref_line(df_eff[y_col]), mode='lines',
                    line=dict(color='#CBD5E1', width=1.5, dash='dash'),
                    name='Referensi Proporsional'
                ))
                marker_kw = dict(
                    size=9,
                    color=df_eff['Gap'] if is_colored else dot_color,
                    colorscale='RdBu' if is_colored else None,
                    cmid=0 if is_colored else None,
                    showscale=is_colored,
                    colorbar=dict(title="Gap", thickness=12, len=0.7,
                                  tickfont=dict(size=10,family='DM Mono, monospace')) if is_colored else None,
                    line=dict(width=1, color='white')
                )
                fig.add_trace(go.Scatter(
                    x=df_eff['Porsi_Miskin'], y=df_eff[y_col],
                    mode='markers', name='Daerah', marker=marker_kw,
                    text=df_eff['Kabupaten_Kota'],
                    hovertemplate="<b>%{text}</b><br>Porsi Miskin: %{x:.4f}<br>Kuota/Kapita: %{y:,.2f}<extra></extra>"
                ))
                fig.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(t=10,l=0,r=0,b=0), height=320, font=CHART_FONT,
                    xaxis=dict(showgrid=True, gridcolor='#F1F5F9', zeroline=False,
                               title="Porsi Kemiskinan", title_font=dict(size=11),
                               tickfont=dict(size=10,family='DM Mono, monospace')),
                    yaxis=dict(showgrid=True, gridcolor='#F1F5F9', zeroline=False,
                               title="Kuota / Kapita Miskin", title_font=dict(size=11),
                               tickfont=dict(size=10,family='DM Mono, monospace')),
                    legend=dict(font=dict(size=11), orientation="h", y=-0.2)
                )
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})

        with col_e1:
            _scatter('Kuota_Per_Kapita', 'Eksisting — Proporsi vs Kuota per Kapita',
                     'Outlier menunjukkan ketidaksesuaian distribusi historis', is_colored=True)
        with col_e2:
            _scatter('Rekomendasi_Per_Kapita', 'Rekomendasi — Proporsi vs Kuota per Kapita',
                     'Distribusi setelah optimasi proporsional diterapkan', dot_color='#1B6CA8')

        # Statistik ringkas
        st.markdown("---")
        st.markdown("##### Ringkasan Statistik Efisiensi")
        cv_eks = df_eff['Kuota_Per_Kapita'].std()       / df_eff['Kuota_Per_Kapita'].mean() * 100
        cv_rek = df_eff['Rekomendasi_Per_Kapita'].std() / df_eff['Rekomendasi_Per_Kapita'].mean() * 100
        n_over  = (df['Gap'] < 0).sum()
        n_under = (df['Gap'] > 0).sum()

        c1, c2, c3, c4 = st.columns(4, gap="small")
        with c1: st.metric("CV Kuota Eksisting",    f"{cv_eks:.1f}%", help="Koefisien variasi — semakin kecil semakin merata")
        with c2: st.metric("CV Setelah Optimasi",   f"{cv_rek:.1f}%", delta=f"{cv_rek-cv_eks:.1f}%", delta_color="inverse")
        with c3: st.metric("Daerah Overfunded",     f"{n_over} kab/kota")
        with c4: st.metric("Daerah Underfunded",    f"{n_under} kab/kota")


# ==============================================================================
# HALAMAN 2 — METODOLOGI
# ==============================================================================
elif menu == "Metodologi":

    st.markdown('<span class="header-tag">Metodologi</span>', unsafe_allow_html=True)
    st.title("Pemodelan Matematika & Simulasi")
    st.markdown('<p class="page-subtitle">Dokumentasi algoritma Linear Programming yang menjadi dasar rekomendasi sistem</p>', unsafe_allow_html=True)

    col_m1, col_m2 = st.columns(2, gap="medium")
    with col_m1:
        with st.container(border=True):
            st.markdown('<p class="card-title">Fungsi Alokasi Proporsional</p>'
                        '<p class="card-subtitle">Menghitung jatah ideal berdasarkan proporsi kemiskinan tanpa intervensi batasan</p>',
                        unsafe_allow_html=True)
            st.latex(r"A_i = w_i \times P")
            st.markdown("""
            | Notasi | Deskripsi |
            |--------|-----------|
            | $A_i$ | Alokasi ideal murni daerah $i$ |
            | $w_i$ | Rasio penduduk miskin daerah $i$ terhadap total Jatim |
            | $P$ | Total kuota pagu provinsi |
            """)
    with col_m2:
        with st.container(border=True):
            st.markdown('<p class="card-title">Fungsi Kendala (Shock Prevention)</p>'
                        '<p class="card-subtitle">Menjaga stabilitas kuota menggunakan batas toleransi perubahan kebijakan</p>',
                        unsafe_allow_html=True)
            st.latex(r"L_i \leq R_i \leq U_i")
            st.markdown("""
            | Notasi | Deskripsi |
            |--------|-----------|
            | $R_i$ | Alokasi rekomendasi final |
            | $L_i$ | Batas bawah: Kuota Awal × (1 − τ) |
            | $U_i$ | Batas atas: Kuota Awal × (1 + τ) |
            """)

    st.markdown("---")
    st.markdown("#### Simulasi Komputasi Per Daerah")
    st.caption("Pilih wilayah untuk menelusuri proses kalkulasi secara langkah demi langkah.")

    simulasi_kota = st.selectbox("Kabupaten / Kota", df['Kabupaten_Kota'].unique())
    ds = df[df['Kabupaten_Kota'] == simulasi_kota].iloc[0]
    w_i = ds['Porsi_Miskin']; a_i = ds['Alokasi_Murni']; e_i = ds[pilihan_bansos]
    l_i = ds['Batas_Bawah'];  u_i = ds['Batas_Atas'];    r_i = ds['Alokasi_Rekomendasi']

    with st.container(border=True):
        st.markdown("**Langkah 1 — Alokasi Proporsional Murni**")
        st.latex(rf"A_{{\text{{{simulasi_kota}}}}} = {w_i:.4f} \times {total_pagu:,.0f} = {a_i:,.0f} \text{{ unit}}")

        st.markdown(f"**Langkah 2 — Batas Toleransi ±{toleransi_persen}% dari Kuota Eksisting ({e_i:,.0f} unit)**")
        st.latex(rf"L = {e_i:,.0f} \times (1 - {toleransi_persen/100}) = {l_i:,.0f} \text{{ unit}}")
        st.latex(rf"U = {e_i:,.0f} \times (1 + {toleransi_persen/100}) = {u_i:,.0f} \text{{ unit}}")

        st.markdown("**Langkah 3 — Keputusan Final (Bounded Optimization)**")

        if r_i == l_i:
            kt = f"Alokasi murni ({a_i:,.0f}) berada di bawah batas bawah. Rekomendasi ditetapkan pada batas bawah: **{r_i:,.0f} unit**."
            bc, bl = "#DC2626", "Terbatas — Batas Bawah"
        elif r_i == u_i:
            kt = f"Alokasi murni ({a_i:,.0f}) melampaui batas atas. Rekomendasi dibatasi pada batas atas: **{r_i:,.0f} unit**."
            bc, bl = "#D97706", "Terbatas — Batas Atas"
        else:
            kt = f"Alokasi murni ({a_i:,.0f}) berada dalam zona toleransi. Rekomendasi mengikuti alokasi proporsional: **{r_i:,.0f} unit**."
            bc, bl = "#059669", "Optimal — Dalam Batas"

        st.markdown(f'<span style="display:inline-block;background:{bc}22;color:{bc};font-size:11px;font-weight:700;padding:3px 10px;border-radius:20px;letter-spacing:.06em;margin-bottom:.75rem;">{bl}</span>', unsafe_allow_html=True)
        st.markdown(kt)

    st.markdown("---")
    st.markdown("#### Visualisasi Bounded Knapsack Tracker")
    st.caption("Grafik ini menunjukkan posisi alokasi final relatif terhadap zona toleransi kebijakan.")

    max_range = max(u_i, a_i) * 1.25
    fig_bullet = go.Figure(go.Indicator(
        mode="number+gauge", value=r_i,
        number={'valueformat':",.0f",'suffix':" unit",'font':{'family':'DM Mono, monospace','size':22,'color':'#0F172A'}},
        domain={'x':[0,1],'y':[0,1]},
        title={'text':f"<b style='font-family:DM Sans;'>{simulasi_kota}</b><br><br><span style='font-size:11px;color:#64748B;font-family:DM Sans;'>Alokasi Rekomendasi Final</span>"},
        gauge={
            'shape':"bullet",
            'axis':{'range':[0,max_range],'tickformat':",.0f",'tickfont':{'family':'DM Mono, monospace','size':10,'color':'#94A3B8'}},
            'threshold':{'line':{'color':'#374151','width':3},'thickness':0.75,'value':a_i},
            'steps':[{'range':[0,l_i],'color':'#FEE2E2'},{'range':[l_i,u_i],'color':'#D1FAE5'},{'range':[u_i,max_range],'color':'#FEE2E2'}],
            'bar':{'color':'#1B6CA8','thickness':0.35}
        }
    ))
    fig_bullet.update_layout(height=180, margin=dict(t=30,b=20,l=160,r=40),
                              paper_bgcolor='rgba(0,0,0,0)', font=dict(family='DM Sans, sans-serif'))
    st.plotly_chart(fig_bullet, use_container_width=True, config={'displayModeBar': False})

    st.markdown("""
    <p style="font-size:12px;color:#94A3B8;margin-top:-0.5rem;">
    <b style="color:#374151;">Garis vertikal hitam</b> = target proporsional murni (A<sub>i</sub>) &nbsp;|&nbsp;
    <b style="color:#059669;">Area hijau</b> = zona toleransi aman [L<sub>i</sub>, U<sub>i</sub>] &nbsp;|&nbsp;
    <b style="color:#1B6CA8;">Batang biru</b> = rekomendasi final (R<sub>i</sub>)
    </p>""", unsafe_allow_html=True)


# --- FOOTER ---
st.markdown("---")
st.markdown("""
<div style='text-align:center;padding:1rem 0 1.5rem;font-family:DM Sans,sans-serif;'>
    <p style='font-size:12px;color:#94A3B8;margin:0;line-height:1.8;'>
        Decision Support System — Optimasi Alokasi Bantuan Sosial Jawa Timur<br>
        <span style='color:#CBD5E1;'>Dikembangkan oleh Mahasiswa Matematika ITS &bull; Kerja Praktik BPS Kabupaten Bojonegoro &copy; 2026</span>
    </p>
</div>
""", unsafe_allow_html=True)